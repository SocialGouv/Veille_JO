"""Recette automatisée (E7) : compare l'Excel produit à l'Excel CIBLE.

Usage :
    python tests/compare_cible.py <fichier_généré.xlsx> <fichier_cible.xlsx>

Verdict en deux volets :

- **volet strict** (doit être identique, sinon `ÉCART` et code retour 1) : sections
  présentes et leur ordre (titres normalisés : la cible historique dit « Nouvelles
  inscription », le format du 23/07/2026 « Nouvelles inscriptions ») ; par section,
  l'ensemble des produits (comparés sur nom racine) ; puis, pour chaque colonne
  présente DANS LES DEUX fichiers : laboratoire, liste (« et »/« & » équivalents),
  taux, date, cibles des hyperliens Prix/Taux/Lien. Une colonne absente d'un des deux
  fichiers (le Taux, retiré du format le 23/07/2026) n'est pas comparée ;
- **volet revue humaine** (listé, jamais bloquant) : la colonne Indication, affichée
  côte à côte — la cible contient des raccourcis de jugement humain non exigibles d'un
  traitement automatique (« Idem que PRADAXA », indication laissée vide, coquilles).
"""

import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from rapprochement import nom_racine

# Titre affiché → titre canonique : la cible historique (28/05/2026) et le format du
# 23/07/2026 diffèrent au pluriel près — mêmes sections pour la recette.
TITRES_SECTIONS = {
    "Nouvelles inscription": "Nouvelles inscriptions",
    "Nouvelles inscriptions": "Nouvelles inscriptions",
    "Hausse de prix": "Hausse de prix",
    "Hausses de prix": "Hausse de prix",
    "Baisse de prix": "Baisse de prix",
    "Baisses de prix": "Baisse de prix",
    "Modification de libellé": "Modification de libellé",
    "Extensions d'indications": "Extensions d'indications",
    "Radiations": "Radiations",
}
LARGEUR_BALAYAGE = 12  # colonnes scannées (A→L) : couvre le format cible avec de la marge


def _normaliser(valeur) -> str:
    """Chaîne comparable : espaces insécables → espaces, espaces de bord ignorés."""
    if valeur is None:
        return ""
    return str(valeur).replace("\xa0", " ").strip()


def _en_date(valeur) -> date | None:
    if isinstance(valeur, datetime):
        return valeur.date()
    if isinstance(valeur, date):
        return valeur
    return None


def _en_taux(valeur):
    """Valeur de taux comparable : nombre si convertible (0.35 ≡ "0.35"), sinon chaîne."""
    texte = _normaliser(valeur)
    try:
        return round(float(texte.replace(",", ".")), 6)
    except ValueError:
        return texte


def _cellules_maitres(ws) -> dict:
    """(ligne, colonne) → cellule maître, pour chaque cellule couverte par une fusion.

    Depuis l'évolution du 22/07/2026, l'export fusionne verticalement Laboratoire,
    Indication et Liste au sein d'un produit : la valeur (et l'hyperlien) vivent sur la
    cellule haute de la plage, les autres lisent None sans cette résolution.
    """
    maitres = {}
    for plage in ws.merged_cells.ranges:
        maitre = ws.cell(row=plage.min_row, column=plage.min_col)
        for ligne in range(plage.min_row, plage.max_row + 1):
            for colonne in range(plage.min_col, plage.max_col + 1):
                maitres[(ligne, colonne)] = maitre
    return maitres


def charger_structure(chemin: Path) -> list[dict]:
    """Lit un xlsx de veille : liste ordonnée de sections {titre, entetes, lignes}.

    Chaque ligne de données est un dict {libellé d'en-tête → (valeur, cible du lien)}.
    Le balayage est structurel (titres de sections connus, en-têtes à la ligne suivante,
    données jusqu'à la ligne vide) : il tolère les décalages de la cible faite main.
    Les cellules couvertes par une fusion verticale lisent leur cellule maître.
    """
    ws = load_workbook(chemin).worksheets[0]
    maitres = _cellules_maitres(ws)
    sections: list[dict] = []
    section_courante: dict | None = None
    colonnes_courantes: list[tuple[int, str]] = []

    for indice_ligne in range(1, ws.max_row + 1):
        cellules = [maitres.get((indice_ligne, c), ws.cell(row=indice_ligne, column=c))
                    for c in range(1, LARGEUR_BALAYAGE + 1)]
        valeurs = [_normaliser(c.value) for c in cellules]

        titre_trouve = next((TITRES_SECTIONS[v] for v in valeurs if v in TITRES_SECTIONS),
                            None)
        if titre_trouve:
            section_courante = {"titre": titre_trouve, "entetes": [], "lignes": []}
            sections.append(section_courante)
            colonnes_courantes = []
            continue
        if section_courante is None:
            continue
        if not section_courante["entetes"]:
            colonnes_courantes = [(i + 1, v) for i, v in enumerate(valeurs) if v]
            section_courante["entetes"] = [v for _, v in colonnes_courantes]
            continue
        if not any(valeurs):
            section_courante = None   # ligne vide : fin de la section
            continue
        ligne = {}
        for indice_colonne, libelle in colonnes_courantes:
            cellule = maitres.get((indice_ligne, indice_colonne),
                                  ws.cell(row=indice_ligne, column=indice_colonne))
            lien = cellule.hyperlink.target if cellule.hyperlink else None
            ligne[libelle] = (cellule.value, lien)
        section_courante["lignes"].append(ligne)
    return sections


def _produit_racine(ligne: dict) -> str:
    valeur, _ = ligne.get("Produit", (None, None))
    return nom_racine(_normaliser(valeur))


def comparer(chemin_genere: Path, chemin_cible: Path) -> tuple[bool, list[str], list[str]]:
    """Compare les deux fichiers : (volet strict conforme ?, écarts, revue humaine)."""
    genere = charger_structure(chemin_genere)
    cible = charger_structure(chemin_cible)
    ecarts: list[str] = []
    revue: list[str] = []

    titres_generes = [s["titre"] for s in genere]
    titres_cible = [s["titre"] for s in cible]
    if titres_generes != titres_cible:
        ecarts.append(f"Sections ou ordre différents : généré {titres_generes} "
                      f"vs cible {titres_cible}")
        return False, ecarts, revue

    for section_generee, section_cible in zip(genere, cible):
        titre = section_cible["titre"]
        # La comparaison se fait par nom racine : depuis le contrat du 23/07/2026 le
        # généré porte une ligne par (nom de médicament, laboratoire) — un même nom
        # peut donc porter plusieurs lignes (génériques multi-laboratoires), chacune
        # devant porter les attributs de la cible.
        lignes_generees: dict[str, list[dict]] = {}
        for l in section_generee["lignes"]:
            lignes_generees.setdefault(_produit_racine(l), []).append(l)
        lignes_cible = {_produit_racine(l): l for l in section_cible["lignes"]}

        if set(lignes_generees) != set(lignes_cible):
            manquants = set(lignes_cible) - set(lignes_generees)
            en_trop = set(lignes_generees) - set(lignes_cible)
            if manquants:
                ecarts.append(f"[{titre}] produits manquants : {sorted(manquants)}")
            if en_trop:
                ecarts.append(f"[{titre}] produits en trop : {sorted(en_trop)}")

        for produit in sorted(set(lignes_generees) & set(lignes_cible)):
            groupe, ligne_cible = lignes_generees[produit], lignes_cible[produit]

            for indice, ligne_generee in enumerate(groupe):
                repere = produit if len(groupe) == 1 else f"{produit} (ligne {indice + 1})"

                for libelle in ("Laboratoire", "Liste"):
                    if libelle in ligne_cible and libelle in ligne_generee:
                        valeur_generee = _normaliser(ligne_generee[libelle][0])
                        valeur_cible = _normaliser(ligne_cible[libelle][0])
                        if libelle == "Liste":
                            # « SS et Collectivité » (cible historique) ≡ « SS &
                            # Collectivité » (format utilisatrice du 23/07/2026).
                            valeur_generee = valeur_generee.replace(" & ", " et ")
                            valeur_cible = valeur_cible.replace(" & ", " et ")
                        if valeur_generee != valeur_cible:
                            ecarts.append(f"[{titre}] {repere} — {libelle} : "
                                          f"« {valeur_generee} » ≠ cible « {valeur_cible} »")

                if "Date" in ligne_cible and "Date" in ligne_generee:
                    date_generee = _en_date(ligne_generee["Date"][0])
                    date_cible = _en_date(ligne_cible["Date"][0])
                    if date_generee != date_cible:
                        ecarts.append(f"[{titre}] {repere} — Date : "
                                      f"{date_generee} ≠ cible {date_cible}")

                # Le Taux a quitté le format le 23/07/2026 : comparé seulement si les
                # DEUX fichiers portent la colonne (recette d'anciens formats).
                if "Taux" in ligne_cible and "Taux" in ligne_generee:
                    taux_genere = _en_taux(ligne_generee["Taux"][0])
                    taux_cible = _en_taux(ligne_cible["Taux"][0])
                    if taux_genere != taux_cible:
                        ecarts.append(f"[{titre}] {repere} — Taux : "
                                      f"{taux_genere!r} ≠ cible {taux_cible!r}")

                # Cellules à hyperlien : seule la CIBLE du lien est comparée (le texte
                # affiché est « Site LégiFrance » depuis le 23/07/2026).
                for libelle in ("Prix", "Taux", "Lien"):
                    if libelle in ligne_cible and libelle in ligne_generee:
                        lien_genere = ligne_generee[libelle][1]
                        lien_cible = ligne_cible[libelle][1]
                        if lien_genere != lien_cible:
                            ecarts.append(f"[{titre}] {repere} — lien {libelle} : "
                                          f"{lien_genere} ≠ cible {lien_cible}")

            if "Indication" in ligne_cible:
                indications_generees = []
                for ligne_generee in groupe:
                    valeur = _normaliser(ligne_generee.get("Indication", (None, None))[0])
                    if valeur not in indications_generees:
                        indications_generees.append(valeur)
                indication_cible = _normaliser(ligne_cible["Indication"][0])
                marque = ("identique" if indications_generees == [indication_cible]
                          else "DIFFÉRENTE")
                rendu_genere = "\n    généré  : ".join(i or "(vide)" for i in indications_generees)
                revue.append(f"[{titre}] {produit} ({marque})\n"
                             f"    cible   : {indication_cible or '(vide)'}\n"
                             f"    généré  : {rendu_genere}")

    return not ecarts, ecarts, revue


def principal(argv: list[str]) -> int:
    if len(argv) != 2:
        print(__doc__)
        return 2
    chemin_genere, chemin_cible = Path(argv[0]), Path(argv[1])
    conforme, ecarts, revue = comparer(chemin_genere, chemin_cible)

    print(f"Généré : {chemin_genere}")
    print(f"Cible  : {chemin_cible}")
    print()
    if conforme:
        print("VOLET STRICT : CONFORME (sections, produits, laboratoires, listes, taux,")
        print("               cibles des hyperliens Prix/Taux, dates)")
    else:
        print("VOLET STRICT : ÉCART")
        for ecart in ecarts:
            print(f"  - {ecart}")
    print()
    print("VOLET REVUE HUMAINE — colonne Indication (jamais bloquant, à faire valider")
    print("par l'utilisatrice : recopie exacte attendue, pas de raccourcis manuels) :")
    for entree in revue:
        print(f"  {entree}")
    return 0 if conforme else 1


if __name__ == "__main__":
    sys.exit(principal(sys.argv[1:]))
