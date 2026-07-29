"""Tests hors ligne d'`export.py` (E6) et de la recette `compare_cible.py` (E7).

Validation « sans clés » (§6 du plan) : la fixture des 8 lignes de l'annexe E est
exportée puis comparée à la CIBLE réelle — volet strict CONFORME attendu. Le fichier
fautif des scripts historiques, lui, doit lever des ÉCARTS (test négatif).
"""

import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
sys.path.insert(0, str(Path(__file__).resolve().parent / "fixtures"))

from openpyxl.workbook import Workbook

import compare_cible
from export import exporter
from fixture_annexe_e import DATE_JO, resultat_consolide_28_05, textes_analyses_28_05
from rapprochement import (SECTION_EXTENSIONS, SECTION_RADIATIONS, LigneConsolidee,
                           ResultatVeille, consolider)

CIBLE = Path(__file__).parent / "fixtures" / "veille_jo_2026-05-28_CIBLE.xlsx"
ANCIEN_EXCEL_FAUTIF = (Path(__file__).resolve().parent.parent.parent
                       / "docs" / "veille_jo_2026-05-28.xlsx")


class TestExportEtRecette(unittest.TestCase):
    def _exporter(self, resultat) -> Path:
        dossier = Path(tempfile.mkdtemp(prefix="veille_test_"))
        return exporter(resultat, dossier=dossier)

    def test_fixture_annexe_e_conforme_a_la_cible(self):
        """E6+E7 : les 8 lignes de l'annexe E, exportées, passent le volet strict."""
        chemin = self._exporter(resultat_consolide_28_05())
        conforme, ecarts, revue = compare_cible.comparer(chemin, CIBLE)
        self.assertTrue(conforme, f"écarts inattendus : {ecarts}")
        self.assertEqual(len(revue), 6)        # 6 indications passées en revue humaine

    def test_chaine_complete_rapprochement_export_recette(self):
        """E5→E6→E7 : les 13 textes analysés, consolidés puis exportés, restent conformes."""
        resultat = consolider(textes_analyses_28_05(), DATE_JO)
        chemin = self._exporter(resultat)
        conforme, ecarts, _revue = compare_cible.comparer(chemin, CIBLE)
        self.assertTrue(conforme, f"écarts inattendus : {ecarts}")

    @unittest.skipUnless(ANCIEN_EXCEL_FAUTIF.exists(),
                         "docs/veille_jo_2026-05-28.xlsx absent (kit de transmission)")
    def test_negatif_ancien_excel_fautif_detecte(self):
        """Le comparateur attrape les pièges historiques (MORPHINE en baisses, ×6, etc.)."""
        conforme, ecarts, _ = compare_cible.comparer(ANCIEN_EXCEL_FAUTIF, CIBLE)
        self.assertFalse(conforme)
        self.assertTrue(ecarts)

    def test_sections_vides_omises_et_titres_exacts(self):
        resultat = resultat_consolide_28_05()
        resultat.lignes = [l for l in resultat.lignes if l.produit == "MORPHINE"]
        chemin = self._exporter(resultat)
        structure = compare_cible.charger_structure(chemin)
        self.assertEqual([s["titre"] for s in structure], ["Hausse de prix"])

    def test_structure_excel_generee(self):
        """Contrat du 23/07/2026 : onglet, colonne A, fusion du titre, en-têtes (mêmes
        colonnes que le mail + Date), plus aucun prix chiffré."""
        from openpyxl import load_workbook
        chemin = self._exporter(resultat_consolide_28_05())
        ws = load_workbook(chemin).worksheets[0]
        self.assertEqual(ws.title, "Veille")
        self.assertEqual(ws.column_dimensions["A"].width, 3)
        self.assertEqual(ws.cell(row=2, column=2).value, "Nouvelles inscriptions")
        self.assertIn("B2:H2", [str(r) for r in ws.merged_cells.ranges])
        self.assertEqual(ws.row_dimensions[2].height, 20)
        # Produit et Laboratoire côte à côte (demande utilisatrice du 22/07 16:48,
        # cas des génériques multi-laboratoires) ; Prix puis Taux en fin de tableau,
        # comme dans la CIBLE.
        self.assertEqual([ws.cell(row=3, column=c).value for c in range(2, 9)],
                         ["Date", "Produit", "Laboratoire", "Indication", "Liste",
                          "Prix", "Taux"])
        # Taux de WEGOVY : nombre au format « 0% » lié à sa décision UNCAM (constat CIBLE).
        cellule_taux = ws.cell(row=4, column=8)
        self.assertEqual(cellule_taux.value, 0.35)
        self.assertEqual(cellule_taux.number_format, "0%")
        self.assertEqual(cellule_taux.hyperlink.target,
                         "https://www.legifrance.gouv.fr/jorf/id/JORFTEXT000054144868")
        # VGENFLI (ligne 6) : aucune décision de taux ne le vise → « N/A » sans lien.
        self.assertEqual(ws.cell(row=6, column=8).value, "N/A")
        self.assertIsNone(ws.cell(row=6, column=8).hyperlink)
        # Ligne WEGOVY : date réelle, produit gras 3A3A3A, labo, liste avec lien vers
        # l'arrêté SS (un seul hyperlien possible par cellule Excel).
        self.assertEqual(ws.cell(row=4, column=2).number_format, "DD/MM/YYYY")
        cellule_produit = ws.cell(row=4, column=3)
        self.assertEqual(cellule_produit.value, "WEGOVY")
        self.assertTrue(cellule_produit.font.bold)
        self.assertEqual(cellule_produit.font.color.rgb, "FF3A3A3A")
        self.assertEqual(ws.cell(row=4, column=4).value, "NOVO NORDISK")
        cellule_liste = ws.cell(row=4, column=6)
        self.assertEqual(cellule_liste.value, "SS & Collectivité")
        self.assertEqual(cellule_liste.hyperlink.target,
                         "https://www.legifrance.gouv.fr/jorf/id/JORFTEXT000054144800")
        # Section Hausse de prix (6 inscriptions : lignes 4-9, vide en 10, titre en 11,
        # en-têtes en 12, MORPHINE en 13) : lien Prix « Site LégiFrance » sans montant.
        self.assertEqual(ws.cell(row=11, column=2).value, "Hausse de prix")
        self.assertEqual([ws.cell(row=12, column=c).value for c in range(2, 6)],
                         ["Date", "Produit", "Laboratoire", "Prix"])
        cellule_prix = ws.cell(row=13, column=5)
        self.assertEqual(cellule_prix.value, "Site LégiFrance")
        self.assertEqual(cellule_prix.hyperlink.target,
                         "https://www.legifrance.gouv.fr/jorf/id/JORFTEXT000054144874")

    def test_permission_error_suffixe_horaire(self):
        """Fichier ouvert dans Excel : l'export bascule sur un nom suffixé de l'heure."""
        dossier = Path(tempfile.mkdtemp(prefix="veille_test_"))
        vrai_save = Workbook.save
        appels = {"n": 0}

        def save_bloque_puis_passe(self, chemin):
            appels["n"] += 1
            if appels["n"] == 1:
                raise PermissionError("fichier ouvert dans Excel")
            return vrai_save(self, chemin)

        with mock.patch.object(Workbook, "save", save_bloque_puis_passe):
            chemin = exporter(resultat_consolide_28_05(), dossier=dossier)
        self.assertTrue(chemin.exists())
        self.assertRegex(chemin.name, r"veille_jo_2026-05-28_\d{6}\.xlsx")


class TestReglesDeSection(unittest.TestCase):
    """Les règles de présentation portées par la ligne consolidée (segments de liste,
    lien de section, rappels SIRTURO) rendues côté Excel — la formulation y diffère
    volontairement de celle du mail : une cellule Excel ne porte qu'un hyperlien, déjà
    pris par la colonne Liste ou Prix, donc le lien de prix est « dans le mail »."""

    URL = "https://www.legifrance.gouv.fr/jorf/id/{}"

    def _feuille(self, lignes):
        from openpyxl import load_workbook
        dossier = Path(tempfile.mkdtemp(prefix="veille_test_"))
        chemin = exporter(ResultatVeille(date_jo=DATE_JO, lignes=lignes), dossier=dossier)
        return load_workbook(chemin).worksheets[0]

    def test_rappels_sirturo_dans_l_indication(self):
        ligne = LigneConsolidee(
            produit="SIRTURO", date_jo=DATE_JO, laboratoire="JANSSEN", racine="SIRTURO",
            indication="En association appropriée…", section=SECTION_EXTENSIONS,
            listes=[("Rétrocession", self.URL.format("JORFTEXT3"))],
            lien_prix=self.URL.format("JORFTEXT4"),
            lien_extension=self.URL.format("JORFTEXT5"))
        ws = self._feuille([ligne])
        self.assertEqual(ws.cell(row=2, column=2).value, "Extensions d'indications")
        cellule_indication = ws.cell(row=4, column=5)
        self.assertEqual(cellule_indication.value,
                         "En association appropriée…\n\nInscription : Rétrocession — "
                         "modification de prix publiée (lien dans le mail)")
        self.assertIsNone(cellule_indication.hyperlink)
        # Colonne Lien : le texte porteur de la section, c'est-à-dire l'extension.
        self.assertEqual(ws.cell(row=4, column=6).hyperlink.target,
                         self.URL.format("JORFTEXT5"))

    def test_radiation_liste_radiee_et_lien_de_l_arrete(self):
        ligne = LigneConsolidee(
            produit="VFEND", date_jo=DATE_JO, laboratoire="PFIZER", racine="VFEND",
            section=SECTION_RADIATIONS,
            listes=[("SS", self.URL.format("JORFTEXT7"))],   # inscription d'un autre jour
            listes_radiation=[("LES MCO", self.URL.format("JORFTEXT6"))])
        ws = self._feuille([ligne])
        self.assertEqual([ws.cell(row=3, column=c).value for c in range(2, 7)],
                         ["Date", "Produit", "Laboratoire", "Liste", "Lien"])
        # La colonne Liste d'une radiation porte les listes RADIÉES, pas les listes
        # d'inscription, et son lien est celui de l'arrêté de radiation.
        self.assertEqual(ws.cell(row=4, column=5).value, "LES MCO")
        self.assertEqual(ws.cell(row=4, column=5).hyperlink.target,
                         self.URL.format("JORFTEXT6"))
        self.assertEqual(ws.cell(row=4, column=6).hyperlink.target,
                         self.URL.format("JORFTEXT6"))


if __name__ == "__main__":
    unittest.main()
