"""Export Excel avec openpyxl, au format du contrat du 23/07/2026.

Mêmes sections et colonnes que le mail (sections : rapprochement.SECTIONS), plus une
colonne Date en tête de chaque tableau (héritage du fichier cible historique). Plus
aucun prix chiffré : les cellules Prix et Lien portent un hyperlien
« Site LégiFrance ». Constats conservés du fichier CIBLE historique
(`tests/fixtures/veille_jo_2026-05-28_CIBLE.xlsx`) :

- la **valeur de taux est un nombre** (0.35) au format `0%` portant l'hyperlien de la
  décision UNCAM, et non une chaîne « 35% » — la CIBLE fait foi ;
- le tableau démarre en **B2** (ligne 1 et colonne A vides, marge du gabarit) ;
- les largeurs de colonnes sont **par colonne physique de la feuille** : on prend, pour
  chaque colonne, le maximum des largeurs des sections présentes.

Une seule cellule Excel ne porte qu'un hyperlien : la colonne Liste affiche tous les
segments (« SS & Collectivité ») mais ne peut lier que le premier arrêté — les liens
par segment sont dans le mail.
"""

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config
from rapprochement import (SECTION_BAISSES, SECTION_EXTENSIONS, SECTION_HAUSSES,
                           SECTION_INSCRIPTIONS, SECTION_MODIFICATIONS,
                           SECTION_RADIATIONS, SECTIONS, LigneConsolidee,
                           RappelExtension, ResultatVeille)

JOURNAL = logging.getLogger("veille_jo.export")

COULEUR_BORDURE_DONNEES = "BFBFBF"
COULEUR_PRODUIT = "3A3A3A"
COULEUR_LIEN = "0000FF"
TEXTE_LIEN = "Site LégiFrance"

COLONNE_DEPART = 2   # colonne B (A vide, largeur 3)
LIGNE_DEPART = 2     # ligne 2 (ligne 1 vide, constat cible)

# Colonnes (libellé, largeur, rendu) de chaque section : les mêmes que le mail, plus la
# colonne Date en tête (héritage du fichier cible historique). Le tableau Radiations n'a
# pas d'exemple dans les mails de l'utilisatrice : colonnes proposées, à confirmer à la
# première occurrence.
COLONNES_EXCEL = {
    SECTION_INSCRIPTIONS: [("Date", 12, "date"), ("Produit", 35, "produit"),
                           ("Laboratoire", 18, "laboratoire"),
                           ("Indication", 80, "indication"), ("Liste", 30, "liste"),
                           ("Prix", 25, "prix"), ("Taux", 10, "taux")],
    SECTION_HAUSSES: [("Date", 12, "date"), ("Produit", 35, "produit"),
                      ("Laboratoire", 18, "laboratoire"), ("Prix", 25, "prix")],
    SECTION_BAISSES: [("Date", 12, "date"), ("Produit", 35, "produit"),
                      ("Laboratoire", 18, "laboratoire"), ("Prix", 25, "prix")],
    SECTION_MODIFICATIONS: [("Date", 12, "date"), ("Produit", 35, "produit"),
                            ("Laboratoire", 18, "laboratoire"), ("Lien", 25, "lien")],
    SECTION_EXTENSIONS: [("Date", 12, "date"), ("Produit", 35, "produit"),
                         ("Laboratoire", 18, "laboratoire"),
                         ("Indication", 80, "indication"), ("Lien", 25, "lien")],
    SECTION_RADIATIONS: [("Date", 12, "date"), ("Produit", 35, "produit"),
                         ("Laboratoire", 18, "laboratoire"), ("Liste", 30, "liste"),
                         ("Lien", 25, "lien")],
}

# (clé de section, titre, couleur de fond, colonnes) — clés, titres, couleurs et ordre
# d'affichage viennent de la source unique `rapprochement.SECTIONS` (openpyxl attend
# l'hexadécimal nu, sans « # »).
SECTIONS_EXCEL = [(cle, titre, couleur, COLONNES_EXCEL[cle])
                  for cle, titre, couleur in SECTIONS]

BORDURE_DONNEES = Border(*[Side(style="thin", color=COULEUR_BORDURE_DONNEES)] * 4)
BORDURE_TITRE = Border(*[Side(style="thin", color="000000")] * 4)
BORDURE_ENTETE = Border(bottom=Side(style="thin", color="000000"))


def _police(taille: int = 11, gras: bool = False, couleur: str | None = None,
            souligne: bool = False) -> Font:
    couleur_argb = couleur
    if couleur and len(couleur) == 6:
        couleur_argb = "FF" + couleur   # alpha opaque explicite (comme la cible)
    return Font(name=config.POLICE, size=taille, bold=gras, color=couleur_argb,
                underline="single" if souligne else None)


def _suffixe_verification(ligne: LigneConsolidee, texte: str) -> str:
    """Suffixe « (à vérifier) » du §5.1 sur le produit d'une ligne douteuse."""
    return f"{texte} (à vérifier)" if ligne.a_verifier else texte


def _texte_rappel(rappel: RappelExtension) -> str:
    """Un rappel SIRTURO en texte brut : les segments de liste sont énumérés, le lien de
    prix renvoie au mail (une cellule Excel ne porte qu'un hyperlien, déjà utilisé)."""
    if rappel.segments:
        return f"{rappel.etiquette} : " + " & ".join(libelle for libelle, _ in rappel.segments)
    return f"{rappel.etiquette} publiée (lien dans le mail)"


def _indication_affichee(ligne: LigneConsolidee) -> str:
    """Indication, suivie des rappels de la ligne le cas échéant (règle SIRTURO)."""
    rappels = ligne.rappels_extension
    if not rappels:
        return ligne.indication
    separateur = "\n\n" if ligne.indication else ""
    return ligne.indication + separateur + " — ".join(_texte_rappel(r) for r in rappels)


def _rendre_date(cellule, ligne: LigneConsolidee) -> None:
    cellule.value = datetime(ligne.date_jo.year, ligne.date_jo.month, ligne.date_jo.day)
    cellule.number_format = "DD/MM/YYYY"
    cellule.alignment = Alignment(horizontal="left", vertical="center")


def _rendre_produit(cellule, ligne: LigneConsolidee) -> None:
    cellule.value = _suffixe_verification(ligne, ligne.produit)
    cellule.font = _police(gras=True, couleur=COULEUR_PRODUIT)
    cellule.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _rendre_laboratoire(cellule, ligne: LigneConsolidee) -> None:
    """Laboratoire, ou « ancien → nouveau » sur un transfert d'exploitation du jour."""
    cellule.value = ligne.laboratoire_affiche


def _rendre_indication(cellule, ligne: LigneConsolidee) -> None:
    cellule.value = _indication_affichee(ligne) or None
    cellule.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _rendre_liste(cellule, ligne: LigneConsolidee) -> None:
    segments = ligne.segments_liste
    cellule.value = " & ".join(libelle for libelle, _ in segments) if segments else "N/A"
    # Un seul hyperlien possible par cellule Excel (contrairement au mail, où chaque
    # segment est cliquable) : le premier arrêté disponible.
    lien_liste = next((lien for _, lien in segments if lien), None)
    if segments and lien_liste:
        cellule.hyperlink = lien_liste
        cellule.font = _police(couleur=COULEUR_LIEN, souligne=True)


def _rendre_lien_texte(cellule, url: str | None) -> None:
    """Cellule « Site LégiFrance » cliquable, « N/A » à défaut d'URL."""
    if url:
        cellule.value = TEXTE_LIEN
        cellule.hyperlink = url
        cellule.font = _police(couleur=COULEUR_LIEN, souligne=True)
    else:
        cellule.value = "N/A"


def _rendre_prix(cellule, ligne: LigneConsolidee) -> None:
    # Décision utilisatrice du 23/07/2026 : lien seul, plus jamais de montant.
    _rendre_lien_texte(cellule, ligne.lien_prix)


def _rendre_taux(cellule, ligne: LigneConsolidee) -> None:
    """Taux de participation : nombre au format « 0% » portant l'hyperlien de la décision
    UNCAM (constat CIBLE, qui fait foi), « N/A » quand aucune décision ne le publie."""
    if ligne.taux == "N/A":
        cellule.value = "N/A"
        return
    cellule.value = float(ligne.taux)
    cellule.number_format = "0%"
    if ligne.lien_taux:
        cellule.hyperlink = ligne.lien_taux
        cellule.font = _police(couleur=COULEUR_LIEN, souligne=True)


def _rendre_lien(cellule, ligne: LigneConsolidee) -> None:
    _rendre_lien_texte(cellule, ligne.lien_section)


# Rendu de chaque étiquette de colonne des tables `COLONNES_EXCEL`. Ajouter une colonne
# se fait ici et dans sa table de section, sans toucher `_remplir_cellule`.
RENDUS_CELLULE = {
    "date": _rendre_date,
    "produit": _rendre_produit,
    "laboratoire": _rendre_laboratoire,
    "indication": _rendre_indication,
    "liste": _rendre_liste,
    "prix": _rendre_prix,
    "taux": _rendre_taux,
    "lien": _rendre_lien,
}


def _remplir_cellule(cellule, rendu: str, ligne: LigneConsolidee) -> None:
    """Remplit et style une cellule de données selon son type de rendu.

    Étiquette de rendu inconnue → `KeyError` : déclarer une colonne sans écrire son
    rendu ne doit jamais produire une cellule vide en silence.
    """
    cellule.border = BORDURE_DONNEES
    cellule.font = _police()
    cellule.alignment = Alignment(horizontal="center", vertical="center")
    RENDUS_CELLULE[rendu](cellule, ligne)


def _dessiner_section(ws, curseur: int, titre: str, couleur: str, colonnes, lignes) -> int:
    """Dessine une section complète (titre fusionné, en-têtes, données). Retourne la
    ligne suivante disponible (après la ligne vide de séparation)."""
    fin_fusion = COLONNE_DEPART + len(colonnes) - 1
    ws.merge_cells(start_row=curseur, start_column=COLONNE_DEPART,
                   end_row=curseur, end_column=fin_fusion)
    cellule_titre = ws.cell(row=curseur, column=COLONNE_DEPART, value=titre)
    cellule_titre.fill = PatternFill(fill_type="solid", start_color=couleur, end_color=couleur)
    cellule_titre.font = _police(taille=12, gras=True)
    cellule_titre.alignment = Alignment(horizontal="center", vertical="center")
    for decalage in range(len(colonnes)):
        ws.cell(row=curseur, column=COLONNE_DEPART + decalage).border = BORDURE_TITRE
    ws.row_dimensions[curseur].height = 20
    curseur += 1

    for decalage, (libelle, _largeur, _rendu) in enumerate(colonnes):
        cellule = ws.cell(row=curseur, column=COLONNE_DEPART + decalage, value=libelle)
        cellule.font = _police(gras=True)
        cellule.alignment = Alignment(horizontal="center", vertical="center")
        cellule.border = BORDURE_ENTETE
    curseur += 1

    for i, ligne in enumerate(lignes):
        for decalage, (_libelle, _largeur, rendu) in enumerate(colonnes):
            _remplir_cellule(ws.cell(row=curseur + i, column=COLONNE_DEPART + decalage),
                             rendu, ligne)
    curseur += len(lignes)
    return curseur + 1  # une ligne vide entre sections (§5.1)


def _appliquer_largeurs(ws, sections_presentes) -> None:
    """Largeur par colonne physique : maximum des largeurs des sections présentes."""
    ws.column_dimensions["A"].width = 3
    largeurs: dict[int, float] = {}
    for _cle, _titre, _couleur, colonnes in sections_presentes:
        for decalage, (_libelle, largeur, _rendu) in enumerate(colonnes):
            indice = COLONNE_DEPART + decalage
            largeurs[indice] = max(largeurs.get(indice, 0), largeur)
    for indice, largeur in largeurs.items():
        ws.column_dimensions[ws.cell(row=1, column=indice).column_letter].width = largeur


def exporter(resultat: ResultatVeille, dossier: Path | None = None) -> Path:
    """Écrit `sorties/veille_jo_AAAA-MM-JJ.xlsx` et retourne son chemin.

    Sections vides omises. Fichier cible ouvert dans Excel (`PermissionError`) :
    repli automatique sur un nom suffixé de l'heure.
    """
    dossier = dossier or Path(__file__).parent / config.DOSSIER_SORTIES
    dossier.mkdir(exist_ok=True)

    classeur = Workbook()
    ws = classeur.active
    ws.title = "Veille"

    sections_presentes = [(cle, titre, couleur, colonnes)
                          for cle, titre, couleur, colonnes in SECTIONS_EXCEL
                          if resultat.lignes_par_section(cle)]
    _appliquer_largeurs(ws, sections_presentes)

    curseur = LIGNE_DEPART
    for cle, titre, couleur, colonnes in sections_presentes:
        lignes = resultat.lignes_par_section(cle)
        curseur = _dessiner_section(ws, curseur, titre, couleur, colonnes, lignes)
        JOURNAL.info("Section « %s » : %d ligne(s).", titre, len(lignes))

    chemin = dossier / f"veille_jo_{resultat.date_jo:%Y-%m-%d}.xlsx"
    try:
        classeur.save(chemin)
    except PermissionError:
        # Fichier ouvert dans Excel : on n'écrase pas, on suffixe l'heure.
        chemin = dossier / f"veille_jo_{resultat.date_jo:%Y-%m-%d}_{datetime.now():%H%M%S}.xlsx"
        classeur.save(chemin)
        JOURNAL.warning("Fichier cible ouvert : export sous %s.", chemin.name)
    JOURNAL.info("Excel écrit : %s", chemin)
    return chemin
